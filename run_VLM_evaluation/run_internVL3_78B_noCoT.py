import math
import os
import shutil
import time
import torch
import numpy as np
import openpyxl
import re
import torchvision.transforms as T
from PIL import Image
from decord import VideoReader, cpu
from transformers import AutoConfig, AutoModel, AutoTokenizer
from jinja2 import Environment, FileSystemLoader

# 瀵煎叆浣犵殑鏈湴宸ュ叿
from utils import read_xlsx_to_list, get_choice_list, is_multiple_choice_question, is_verified_correct, get_gold_answer, apply_schema_header_aliases, get_question_text
from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths

# ==========================================
# 1. 鏍稿績宸ュ叿鍑芥暟
# ==========================================

def parse_thought_and_answer(response: str):
    """瑙ｆ瀽 <think> 鍜?<answer> 鏍囩锛屽苟鍋?A-D 鍏滃簳"""
    thought, answer = "", ""
    thought_match = re.search(r'<think>(.*?)</think>', response, re.DOTALL)
    if thought_match:
        thought = thought_match.group(1).strip()
    
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.DOTALL)
    if answer_match:
        answer = answer_match.group(1).strip()
    
    # 鍏滃簳閫昏緫锛氬鏋滄病鎵惧埌鏍囩锛屾彁鍙栨渶鍚庝竴涓嚭鐜扮殑 A-D 瀛楁瘝
    if not answer:
        potential = re.findall(r'\b([A-D])\b', response)
        answer = potential[-1] if potential else response
    return thought, answer

def get_video_tensors_fast(video_path: str, input_size: int = 448):
    """48 甯ф娊甯х瓥鐣ワ細纭繚 Benchmark 姣旇緝鐨勫叕骞虫€?"""
    vr = VideoReader(video_path, ctx=cpu(0))
    fps = vr.get_avg_fps()
    duration = len(vr) / fps
    # 绛栫暐锛?=48s 鎸?1fps锛?48s 鎶?48 甯?
    num_segments = max(1, int(round(duration))) if duration <= 48.0 else 48
    frame_indices = np.linspace(0, len(vr) - 1, num_segments, dtype=int)
    
    frames_np = vr.get_batch(frame_indices).asnumpy()
    frames_tensor = torch.from_numpy(frames_np).permute(0, 3, 1, 2).float() / 255.0
    
    transform = T.Compose([
        T.Resize((input_size, input_size), interpolation=T.InterpolationMode.BICUBIC, antialias=True),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225))
    ])
    return transform(frames_tensor)


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_item_key(item):
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"
def build_completed_question_set(sheet, header_index, answer_col):
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        predicted_text = str(predicted).strip() if predicted is not None else ""
        if predicted_text and not predicted_text.startswith("[SKIP]") and "ERROR" not in predicted_text and predicted_text != "[ERR]":
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value,
                "question": sheet.cell(row=row, column=header_index["question"]).value,
            }
            completed.add(build_item_key(row_item))
    return completed


def build_question_row_map(sheet, header_index):
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value,
            "question": sheet.cell(row=row, column=header_index["question"]).value,
        }
        question_row_map[build_item_key(row_item)] = row
    return question_row_map


def ensure_result_row(sheet, header_index, question_row_map, item):
    item_key = build_item_key(item)
    if item_key in question_row_map:
        return question_row_map[item_key]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Version", "Reasoning_Log", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    question_row_map[item_key] = row
    return row


def ensure_single_process_launch(model_name: str):
    world_size = int(os.environ.get("WORLD_SIZE", "1"))
    if world_size <= 1:
        return
    local_rank = os.environ.get("LOCAL_RANK", "0")
    raise RuntimeError(
        f"{model_name} must run in a single process. WORLD_SIZE={world_size}, LOCAL_RANK={local_rank}"
    )
def split_model(model_path: str):
    """
    鎸?InternVL 瀹樻柟寤鸿鍒囧垎 78B锛屽鍗℃椂淇濊瘉棣栧熬鍏抽敭灞備笌瑙嗚鍒嗘敮钀藉湪鍚屼竴寮犲崱锛?
    閬垮厤 chat 鎺ㄧ悊杩囩▼涓嚭鐜拌法璁惧寮犻噺閿欒銆?
    """
    world_size = torch.cuda.device_count()
    if world_size < 3:
        raise RuntimeError("Invalid runtime state")

    config = AutoConfig.from_pretrained(model_path, trust_remote_code=True)
    num_layers = config.llm_config.num_hidden_layers
    device_map = {}

    num_layers_per_gpu = math.ceil(num_layers / (world_size - 0.5))
    num_layers_per_gpu = [num_layers_per_gpu] * world_size
    num_layers_per_gpu[0] = math.ceil(num_layers_per_gpu[0] * 0.5)

    layer_cnt = 0
    for i, num_layer in enumerate(num_layers_per_gpu):
        for _ in range(num_layer):
            if layer_cnt >= num_layers:
                break
            device_map[f"language_model.model.layers.{layer_cnt}"] = i
            layer_cnt += 1

    device_map["vision_model"] = 0
    device_map["mlp1"] = 0
    device_map["language_model.model.tok_embeddings"] = 0
    device_map["language_model.model.embed_tokens"] = 0
    device_map["language_model.output"] = 0
    device_map["language_model.model.norm"] = 0
    device_map["language_model.model.rotary_emb"] = 0
    device_map["language_model.lm_head"] = 0
    device_map[f"language_model.model.layers.{num_layers - 1}"] = 0
    return device_map

# ==========================================
# 2. 涓绘帹鐞嗘祦绋?
# ==========================================
if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_PATH = "OpenGVLab/InternVL3-78B"
    MODEL_NAME = "InternVL3-78B-NoCoT"
    
    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    LANGUAGE = 'EN'
    TEMPLATE_PATH = "templates/qa_prompt_EN_noCoT.md"

    # --- 1. 鍔犺浇妯″瀷 (閽堝澶氬崱鐜浼樺寲) ---
    ensure_single_process_launch(MODEL_NAME)
    print(f"馃専 姝ｅ湪鍒濆鍖栨ā鍨? {MODEL_NAME}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH, trust_remote_code=True, use_fast=False)
    device_map = split_model(MODEL_PATH)
    
    model = AutoModel.from_pretrained(
        MODEL_PATH,
        torch_dtype=torch.bfloat16,
        trust_remote_code=True,
        device_map=device_map,
        use_flash_attn=True,   # 姝ゆ椂浣犵殑 FA2 搴旇宸茬粡瀹岀編宸ヤ綔
        low_cpu_mem_usage=True # 閬垮厤 meta tensor 鎶ラ敊
    ).eval()

    # --- 2. Excel 鍒濆鍖栦笌鏂偣缁紶 ---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    
    if os.path.exists(result_path):
        print(f"馃搨 鍙戠幇宸叉湁瀛樻。锛屾鍦ㄦ墽琛屾柇鐐圭画浼?..")
        wb = openpyxl.load_workbook(result_path)
    else:
        print(f"馃啎 鍒涘缓鏂扮粨鏋滄枃浠? {result_path}")
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        # 鍒濆鍖栬〃澶?
        base_col = sheet.max_column
        sheet.cell(1, base_col + 1, "Model_Version")
        sheet.cell(1, base_col + 2, "Reasoning_Log")
        sheet.cell(1, base_col + 3, "Predicted_Answer")
        wb.save(result_path)

    sheet = wb.active
    header_index = build_header_index(sheet)
    col_model = header_index["Model_Version"]
    col_think = header_index["Reasoning_Log"]
    col_ans = header_index["Predicted_Answer"]
    completed_questions = build_completed_question_set(sheet, header_index, col_ans)
    question_row_map = build_question_row_map(sheet, header_index)

    # --- 3. 寰幆鎺ㄧ悊 ---
    gen_config = dict(max_new_tokens=1024, do_sample=False)
    jinja_env = Environment(loader=FileSystemLoader('.'))

    print(f"馃殌 寮€濮嬭窇娴嬶紝鐩爣棰樻暟: {len(data)}")

    for i, item in enumerate(data):
        item_key = build_item_key(item)
        if item_key in completed_questions:
            print("[LOG]")
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)

        # [閫昏緫 2]: 鍩虹杩囨护
        if not is_verified_correct(item): continue

        # [閫昏緫 3]: 鎸夐棶棰樼被鍨嬭烦杩囬潪閫夋嫨棰?
        if not is_multiple_choice_question(item):
            print(f"鈴?璺宠繃绗?{i} 棰? 妫€娴嬩负绠€绛旈 (Open-ended)")
            continue

        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        if not video_path or not os.path.exists(video_path):
            sheet.cell(row, col_ans, "[SKIP] File Missing")
            continue

        try:
            # 瑙嗚棰勫鐞?
            pixel_values = get_video_tensors_fast(video_path).to("cuda:0", dtype=torch.bfloat16)
            num_frames = pixel_values.shape[0]
            
            # Prompt 鍑嗗
            q_data = {
                "Question": get_question_text(item, LANGUAGE),
                "Choices": get_choice_list(item)
            }
            video_prefix = ''.join([f'Frame{j+1}: <image>\n' for j in range(num_frames)])
            prompt_text = jinja_env.get_template(TEMPLATE_PATH).render(**q_data)
            
            # 妯″瀷璋冪敤
            response = model.chat(
                tokenizer, pixel_values, video_prefix + prompt_text, gen_config,
                num_patches_list=[1] * num_frames, history=None, return_history=False
            )
            if isinstance(response, tuple): response = response[0]
            
            # 瑙ｆ瀽涓庡瓨鍌?
            thought, answer = parse_thought_and_answer(response)
            
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_think, thought)
            sheet.cell(row, col_ans, answer)
            completed_questions.add(item_key)
            
            print(f"鉁?[{i}/{len(data)}] 棰勬祴: {answer} | 姝ｇ‘: {get_gold_answer(item, LANGUAGE)}")
            
            # 姣?5 棰樺瓨鐩樹竴娆?
            if i % 5 == 0: wb.save(result_path)

        except Exception as e:
            print(f"鉂?绗?{i} 棰樺嚭閿? {e}")
            sheet.cell(row, col_ans, f"ERROR: {str(e)}")
            wb.save(result_path)
            torch.cuda.empty_cache()

    wb.save(result_path)
    print(f"\n馃弫 鎺ㄧ悊浠诲姟鍦嗘弧缁撴潫锛佹渶缁堝瓨妗? {result_path}")




