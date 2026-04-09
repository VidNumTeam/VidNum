import os
import shutil
import time
import torch
import numpy as np
import openpyxl
import re
import torchvision.transforms as T
from PIL import Image
from decord import VideoReader, cpu
from transformers import AutoModel, AutoTokenizer
from jinja2 import Environment, FileSystemLoader

# 瀵煎叆浣犲師鏈夌殑鏈湴宸ュ叿
from utils import read_xlsx_to_list, get_choice_list, is_multiple_choice_question, is_verified_correct, get_gold_answer, apply_schema_header_aliases, get_question_text
from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths

# ==========================================
# 1. 鏍稿績宸ュ叿鍑芥暟 (淇濈暀浣犵殑绮剧粏鍖栨娊甯у拰杩囨护)
# ==========================================

def parse_thought_and_answer(response: str):
    """绮惧噯瑙ｆ瀽 <think> 鍜?<answer>"""
    thought, answer = "", ""
    thought_match = re.search(r'<think>(.*?)</think>', response, re.DOTALL)
    if thought_match:
        thought = thought_match.group(1).strip()
    
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.DOTALL)
    if answer_match:
        answer = answer_match.group(1).strip()
    
    # 鍏滃簳锛氬鏋滄ā鍨嬭皟鐨病鍔犳爣绛撅紝鎵炬渶鍚庝竴涓嚭鐜扮殑 A-D
    if not answer:
        potential = re.findall(r'\b([A-D])\b', response)
        answer = potential[-1] if potential else response
    return thought, answer

def is_flash_attn_available() -> bool:
    """甯︾‖浠剁増鏈娴嬬殑 FA 妫€鏌?"""
    try:
        import flash_attn
        if torch.cuda.is_available():
            major, _ = torch.cuda.get_device_capability()
            if major >= 8: return True # H100 (9.0) 鑲畾鏀寔
            print(f"鈿狅笍 鏄惧崱鏋舵瀯鐗堟湰 {major} 杩囦綆锛屾棤娉曞惎鐢?FA2")
        return False
    except Exception:
        return False

def get_video_tensors_fast(video_path: str, input_size: int = 448):
    """淇濈暀浣犱箣鍓嶇殑 1fps 鎶藉抚绛栫暐锛屼絾浣跨敤鎵瑰鐞嗗姞閫?"""
    vr = VideoReader(video_path, ctx=cpu(0))
    fps = vr.get_avg_fps()
    duration = len(vr) / fps
    
    # 浣犵殑鏍稿績绛栫暐锛?=48s 鎸?1fps锛?48s 鎶?48 甯?
    num_segments = max(1, int(round(duration))) if duration <= 48.0 else 48
    frame_indices = np.linspace(0, len(vr) - 1, num_segments, dtype=int)
    
    frames_np = vr.get_batch(frame_indices).asnumpy()
    frames_tensor = torch.from_numpy(frames_np).permute(0, 3, 1, 2).float() / 255.0
    
    transform = T.Compose([
        T.Resize((input_size, input_size), interpolation=T.InterpolationMode.BICUBIC, antialias=True),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225))
    ])
    return transform(frames_tensor)


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_item_key(item):
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"
def build_completed_question_set(sheet, header_index, answer_col):
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        predicted_text = str(predicted).strip() if predicted is not None else ""
        if predicted_text and not predicted_text.startswith("[SKIP]") and "ERROR" not in predicted_text and predicted_text != "[ERR]":
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value,
                "question": sheet.cell(row=row, column=header_index["question"]).value,
            }
            completed.add(build_item_key(row_item))
    return completed


def build_question_row_map(sheet, header_index):
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value,
            "question": sheet.cell(row=row, column=header_index["question"]).value,
        }
        question_row_map[build_item_key(row_item)] = row
    return question_row_map


def ensure_result_row(sheet, header_index, question_row_map, item):
    item_key = build_item_key(item)
    if item_key in question_row_map:
        return question_row_map[item_key]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Size", "Thinking_Process", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    question_row_map[item_key] = row
    return row

# ==========================================
# 2. 涓绘帹鐞嗘祦绋?
# ==========================================
if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_PATH = "OpenGVLab/InternVL2_5-8B" 
    MODEL_NAME = f"{MODEL_PATH.split('/')[-1]}-NoCoT"
    
    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    LANGUAGE = 'EN'
    TEMPLATE_PATH = "templates/qa_prompt_EN_noCoT.md"

    # --- 1. 鍔犺浇妯″瀷 ---
    print(f"馃専 寮€濮?Benchmark: {MODEL_NAME} | 璇█: {LANGUAGE}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH, trust_remote_code=True, use_fast=False)
    
    use_fa = is_flash_attn_available()
    print(f"鈿?FlashAttention: {'ON' if use_fa else 'OFF'}")
    
    model = AutoModel.from_pretrained(
        MODEL_PATH, torch_dtype=torch.bfloat16, trust_remote_code=True,
        device_map="auto", use_flash_attn=use_fa
    ).eval()

    # --- 2. Excel 鍒濆鍖?缁窇 ---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    if os.path.exists(result_path):
        wb = openpyxl.load_workbook(result_path)
    else:
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        base_col = sheet.max_column
        sheet.cell(1, base_col + 1, "Model_Size")
        sheet.cell(1, base_col + 2, "Thinking_Process")
        sheet.cell(1, base_col + 3, "Predicted_Answer")
        wb.save(result_path)

    sheet = wb.active
    header_index = build_header_index(sheet)
    col_model = header_index["Model_Size"]
    col_think = header_index["Thinking_Process"]
    col_ans = header_index["Predicted_Answer"]
    completed_questions = build_completed_question_set(sheet, header_index, col_ans)
    question_row_map = build_question_row_map(sheet, header_index)

    # --- 3. 寰幆璺戞祴 ---
    gen_config = dict(max_new_tokens=512, do_sample=False)
    jinja_env = Environment(loader=FileSystemLoader('.'))

    for i, item in enumerate(data):
        item_key = build_item_key(item)
        print(f"\n[{MODEL_NAME}] 姝ｅ湪璺戠 {i} 棰?..")

        if item_key in completed_questions:
            print(f"  > 宸插畬鎴愶紝璺宠繃: {str(get_question_text(item, LANGUAGE if 'LANGUAGE' in globals() else 'EN'))[:30]}...")
            continue
        if not is_multiple_choice_question(item):
            print(f"  > 璺宠繃闈為€夋嫨棰? {str(get_question_text(item, LANGUAGE if 'LANGUAGE' in globals() else 'EN'))[:30]}...")
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)
        
        # --- 涓ユ牸鏁版嵁杩囨护閫昏緫 ---
        if not is_verified_correct(item):
            sheet.cell(row, col_ans, "[SKIP] Check Failed")
            continue
            
        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        if not video_path or not os.path.exists(video_path):
            sheet.cell(row, col_ans, "[SKIP] Video Missing")
            continue

        try:
            # 1. 鍑嗗瑙嗚寮犻噺
            pixel_values = get_video_tensors_fast(video_path).to(model.device, dtype=torch.bfloat16)
            num_frames = pixel_values.shape[0]
            
            # 2. 娓叉煋 Prompt
            q_data = {
                "Question": get_question_text(item, LANGUAGE),
                "Choices": get_choice_list(item)
            }
            video_prefix = ''.join([f'Frame{j+1}: <image>\n' for j in range(num_frames)])
            prompt_text = jinja_env.get_template(TEMPLATE_PATH).render(**q_data)
            
            # 3. 鏍稿績鎺ㄧ悊
            response = model.chat(
                tokenizer, pixel_values, video_prefix + prompt_text, gen_config,
                num_patches_list=[1] * num_frames, history=None, return_history=False
            )
            if isinstance(response, tuple): response = response[0]
            
            # 4. 瑙ｆ瀽涓庢寔涔呭寲
            thought, answer = parse_thought_and_answer(response)
            
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_think, thought)
            sheet.cell(row, col_ans, answer)
            completed_questions.add(item_key)
            
            # 鎵撳嵃涓€涓嬶紝鏂逛究浣犲湪缁堢鐩潃杩涘害
            print(f"  > Answer: {answer} (True: {get_gold_answer(item, LANGUAGE)})")
            
            if i % 5 == 0: wb.save(result_path) # 闂存瓏鎬т繚瀛橈紝绋冲鑰佺嫍

        except Exception as e:
            print(f"  鉂?Error on row {row}: {e}")
            sheet.cell(row, col_ans, f"ERROR: {str(e)}")
            wb.save(result_path)
            if "OutOfMemory" in str(e): torch.cuda.empty_cache()

    wb.save(result_path)
    print(f"\n鉁?鍏ㄩ儴鎼炲畾锛佺粨鏋滃湪: {result_path}")



