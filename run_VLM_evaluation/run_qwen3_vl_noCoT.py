import os
import torch

import openpyxl
import re
import shutil
from decord import VideoReader, cpu
from transformers import AutoProcessor, AutoModelForImageTextToText, Qwen3VLForConditionalGeneration
from qwen_vl_utils import process_vision_info
from jinja2 import Environment, FileSystemLoader
from utils import get_choice_list, is_multiple_choice_question, is_verified_correct, get_gold_answer, apply_schema_header_aliases, get_question_text


def parse_answer(response: str):
    """
    鍚屾椂鎻愬彇鎬濈淮閾?think)鍜屾渶缁堢瓟妗?answer)
    杩斿洖: (, final_answer)
    """
    # 2. 鎻愬彇 <answer> 鏍囩鍐呭
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.IGNORECASE | re.DOTALL)
    
    if answer_match:
        final_answer = answer_match.group(1).strip().upper()
    else:
        # 鍏滃簳閫昏緫锛氬鏋滆鎴柇浜嗘垨鑰呮病鍐欐爣绛撅紝鎵炬枃鏈腑鏈€鍚庝竴涓?A-D
        potential = re.findall(r'\b([A-D])\b', response)
        final_answer = potential[-1].upper() if potential else "[ERR]"

    return final_answer


def get_video_metadata(video_path: str) -> dict:
    vr = VideoReader(video_path, ctx=cpu(0))
    # 鍘熷瑙嗛鐨勫疄闄?FPS
    original_fps = float(vr.get_avg_fps())
    # 蹇呴』浣跨敤 'nframes' 杩欎釜閿悕
    nframes = len(vr)
    duration = nframes / original_fps if original_fps > 0 else 0.0
    return {
        "fps": original_fps,  # 杩欓噷鐨?fps 鏄憡璇夋ā鍨嬪師濮嬭棰戞湁澶氬揩
        "nframes": nframes,   # 蹇呴』鍙?nframes
    }


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_item_key(item):
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"
def build_completed_question_set(sheet, answer_col):
    header_index = build_header_index(sheet)
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        if predicted and not str(predicted).startswith("[SKIP]") and "ERROR" not in str(predicted):
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value if "ID" in header_index else None,
                "question": sheet.cell(row=row, column=header_index["question"]).value if "question" in header_index else None,
            }
            completed.add(build_item_key(row_item))
    return completed
def build_question_row_map(sheet):
    header_index = build_header_index(sheet)
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value if "ID" in header_index else None,
            "question": sheet.cell(row=row, column=header_index["question"]).value if "question" in header_index else None,
        }
        question_row_map[build_item_key(row_item)] = row
    return question_row_map
def ensure_result_row(sheet, header_index, question_row_map, item):
    item_key = build_item_key(item)
    if item_key in question_row_map:
        return question_row_map[item_key]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Version", "Full_Response", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    question_row_map[item_key] = row
    return row


def ensure_result_headers(sheet):
    header_index = build_header_index(sheet)
    required_headers = ["Model_Version", "Full_Response", "Predicted_Answer"]
    missing_headers = [header for header in required_headers if header not in header_index]
    if not missing_headers:
        return build_header_index(sheet)

    next_col = sheet.max_column + 1
    for header in required_headers:
        if header not in header_index:
            sheet.cell(row=1, column=next_col, value=header)
            next_col += 1
    return build_header_index(sheet)

if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_ID = "Qwen/Qwen3-VL-8B-Instruct"
    MODEL_NAME = "Qwen3-VL-8B-NoCoT"
    
    from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths
    from utils import read_xlsx_to_list
    
    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    TEMPLATE_PATH = "templates/qa_prompt_qw3_noCoT.md" # 淇濇寔浣犵殑妯℃澘

    # --- 1. 瀹樻柟鏂瑰紡鍔犺浇 ---
    print(f"馃殌 姝ｅ湪鍔犺浇 {MODEL_NAME}...")
    model = Qwen3VLForConditionalGeneration.from_pretrained(
        MODEL_ID,
        dtype=torch.bfloat16,
        attn_implementation="flash_attention_2",
        device_map="auto", 
    )
    processor = AutoProcessor.from_pretrained(MODEL_ID)

    # --- 2. Excel 鍑嗗 ---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    if os.path.exists(result_path):
        wb = openpyxl.load_workbook(result_path)
    else:
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
    sheet = wb.active
    header_index = ensure_result_headers(sheet)
    wb.save(result_path)
    col_model = header_index["Model_Version"]
    col_full = header_index["Full_Response"]
    col_ans = header_index["Predicted_Answer"]
    completed_questions = build_completed_question_set(sheet, col_ans)
    question_row_map = build_question_row_map(sheet)
    jinja_env = Environment(loader=FileSystemLoader('.'))

    # --- 3. 鏍稿績寰幆 ---
    for i, item in enumerate(data):
        print(f"馃殌 姝ｅ湪澶勭悊绗?{i+1}/{len(data)} 棰?..")
        if not is_multiple_choice_question(item):
            continue
        if not is_verified_correct(item):
            continue

        item_key = build_item_key(item)
        if item_key in completed_questions:
            print("[LOG]")
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)

        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        abs_video_path = os.path.abspath(video_path)
        if not os.path.exists(abs_video_path): continue

        try:
            video_metadata = get_video_metadata(abs_video_path)
            # 娓叉煋浣犵殑鍘熷 Prompt
            q_data = {
                "Question": get_question_text(item, LANGUAGE if "LANGUAGE" in globals() else "EN"),
                "Choices": get_choice_list(item)
            }
            prompt_text = jinja_env.get_template(TEMPLATE_PATH).render(**q_data)

            messages = [{
                "role": "user",
                "content": [
                    {
                        "type": "video",
                        "video": abs_video_path,
                        "fps": 1.0,
                        "video_metadata": video_metadata,
                    },
                    {"type": "text", "text": prompt_text},
                ]
            }]

            # 1. 鏂囨湰棰勫鐞?(鐢熸垚甯︽爣绛剧殑瀛楃涓?
            text = processor.apply_chat_template(messages, add_generation_prompt=True, tokenize=False)
            
            # 2. 瑙嗚鐗瑰緛鎻愬彇
            image_inputs, video_inputs = process_vision_info(messages)
            
            # 3. 缁熶竴缂栫爜
            inputs = processor(
                text=[text],
                images=image_inputs,
                videos=video_inputs,
                padding=True,
                return_tensors="pt",
            ).to(model.device)

          
            # 4. 鎵ц鐢熸垚
            with torch.no_grad():
                output = model.generate(
                    **inputs, 
                    max_new_tokens=512, 
                    do_sample=False,
                    use_cache=True,
                    pad_token_id=processor.tokenizer.pad_token_id
                )
            
            # 5. 瑙ｇ爜涓庝繚瀛?
            input_len = inputs["input_ids"].shape[-1]
            response = processor.decode(output[0][input_len:], skip_special_tokens=True)
            
            answer = parse_answer(response)
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_full, response)
            sheet.cell(row, col_ans, answer)
            completed_questions.add(item_key)
            
            print(f"鉁?[{i+1}] 棰勬祴: {answer}")
            if (i + 1) % 5 == 0: 
                wb.save(result_path)

        except Exception as e:
            print(f"鉂?绱㈠紩 {i} 宕╂簝: {e}")
            import traceback
            traceback.print_exc()
            torch.cuda.empty_cache()

    wb.save(result_path)



