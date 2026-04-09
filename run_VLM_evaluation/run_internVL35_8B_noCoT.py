import os
import shutil
import time
import torch
import numpy as np
import openpyxl
import re
import torchvision.transforms as T
from PIL import Image
from decord import VideoReader, cpu
from transformers import AutoModel, AutoTokenizer
from jinja2 import Environment, FileSystemLoader

# 瀵煎叆浣犵殑鏈湴宸ュ叿
from utils import read_xlsx_to_list, get_choice_list, is_verified_correct, apply_schema_header_aliases, get_question_text
from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths

# ==========================================
# 1. 鏍稿績宸ュ叿鍑芥暟
# ==========================================

def parse_thought_and_answer(response: str):
    """瑙ｆ瀽 <think> 鍜?<answer> 鏍囩"""
    thought, answer = "", ""
    thought_match = re.search(r'<think>(.*?)</think>', response, re.DOTALL)
    if thought_match:
        thought = thought_match.group(1).strip()
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.DOTALL)
    if answer_match:
        answer = answer_match.group(1).strip()
    if not answer:
        potential = re.findall(r'\b([A-D])\b', response)
        answer = potential[-1] if potential else response
    return thought, answer

def get_video_tensors_fast(video_path: str, input_size: int = 448):
    """48 甯ф娊甯х瓥鐣?"""
    vr = VideoReader(video_path, ctx=cpu(0))
    fps = vr.get_avg_fps()
    duration = len(vr) / fps
    num_segments = max(1, int(round(duration))) if duration <= 48.0 else 48
    frame_indices = np.linspace(0, len(vr) - 1, num_segments, dtype=int)
    frames_np = vr.get_batch(frame_indices).asnumpy()
    frames_tensor = torch.from_numpy(frames_np).permute(0, 3, 1, 2).float() / 255.0
    transform = T.Compose([
        T.Resize((input_size, input_size), interpolation=T.InterpolationMode.BICUBIC, antialias=True),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225))
    ])
    return transform(frames_tensor)


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_item_key(item):
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"
def build_completed_question_set(sheet, header_index, answer_col):
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        predicted_text = str(predicted).strip() if predicted is not None else ""
        if predicted_text and not predicted_text.startswith("[SKIP]") and "ERROR" not in predicted_text and predicted_text != "[ERR]":
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value if "ID" in header_index else None,
                "question": sheet.cell(row=row, column=header_index["question"]).value if "question" in header_index else None,
            }
            completed.add(build_item_key(row_item))
    return completed


def build_question_row_map(sheet, header_index):
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value if "ID" in header_index else None,
            "question": sheet.cell(row=row, column=header_index["question"]).value if "question" in header_index else None,
        }
        question_row_map[build_item_key(row_item)] = row
    return question_row_map


def ensure_result_row(sheet, header_index, question_row_map, item):
    item_key = build_item_key(item)
    if item_key in question_row_map:
        return question_row_map[item_key]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Version", "Reasoning_Log", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    question_row_map[item_key] = row
    return row

# ==========================================
# 2. 涓绘帹鐞嗘祦绋?
# ==========================================
if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_PATH = "OpenGVLab/InternVL3_5-8B" 
    MODEL_NAME = "InternVL3_5-8B-NoCoT"
    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    LANGUAGE = 'EN'
    TEMPLATE_PATH = "templates/qa_prompt_EN_noCoT.md"

    # --- 1. 鍔犺浇妯″瀷 (鍗曞崱妯″紡) ---
    print(f"馃専 姝ｅ湪鍔犺浇妯″瀷: {MODEL_NAME}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH, trust_remote_code=True, use_fast=False)
    model = AutoModel.from_pretrained(
        MODEL_PATH,
        torch_dtype=torch.bfloat16,
        trust_remote_code=True,
        device_map="auto",
        use_flash_attn=True,
        low_cpu_mem_usage=False 
    ).eval()

    # --- 2. Excel 鍒濆鍖栦笌鏂偣妫€鏌?---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    
    if os.path.exists(result_path):
        print(f"馃搨 鍙戠幇宸叉湁缁撴灉鏂囦欢锛屽噯澶囩画浼? {result_path}")
        wb = openpyxl.load_workbook(result_path)
    else:
        print(f"馃啎 鍒涘缓鏂扮粨鏋滄枃浠? {result_path}")
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        # 鍒濆鍖栬〃澶?
        base_col = sheet.max_column
        sheet.cell(1, base_col + 1, "Model_Version")
        sheet.cell(1, base_col + 2, "Reasoning_Log")
        sheet.cell(1, base_col + 3, "Predicted_Answer")
        wb.save(result_path)

    sheet = wb.active
    header_index = build_header_index(sheet)
    col_model = header_index["Model_Version"]
    col_think = header_index["Reasoning_Log"]
    col_ans = header_index["Predicted_Answer"]
    completed_questions = build_completed_question_set(sheet, header_index, col_ans)
    question_row_map = build_question_row_map(sheet, header_index)

    # --- 3. 寰幆鎺ㄧ悊 ---
    gen_config = dict(max_new_tokens=1024, do_sample=False)
    jinja_env = Environment(loader=FileSystemLoader('.'))

    for i, item in enumerate(data):
        item_key = build_item_key(item)
        if item_key in completed_questions:
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)

        print(f"馃幀 澶勭悊涓? {i}/{len(data)} | {MODEL_NAME}")
        
        if not is_verified_correct(item): continue
        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        if not video_path or not os.path.exists(video_path): continue

        try:
            pixel_values = get_video_tensors_fast(video_path).to(model.device, dtype=torch.bfloat16)
            num_frames = pixel_values.shape[0]
            
            q_data = {
                "Question": get_question_text(item, LANGUAGE),
                "Choices": get_choice_list(item)
            }
            video_prefix = ''.join([f'Frame{j+1}: <image>\n' for j in range(num_frames)])
            prompt_text = jinja_env.get_template(TEMPLATE_PATH).render(**q_data)
            
            response = model.chat(
                tokenizer, pixel_values, video_prefix + prompt_text, gen_config,
                num_patches_list=[1] * num_frames, history=None, return_history=False
            )
            if isinstance(response, tuple): response = response[0]
            
            thought, answer = parse_thought_and_answer(response)
            
            # 鍐欏叆缁撴灉
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_think, thought)
            sheet.cell(row, col_ans, answer)
            completed_questions.add(item_key)
            
            print(f"  > [棰勬祴]: {answer}")
            
            # 姣?5 棰樹繚瀛樹竴娆★紝鍏奸【瀹夊叏涓庨€熷害
            if i % 5 == 0: wb.save(result_path)

        except Exception as e:
            print(f"  鉂?鍑洪敊: {e}")
            sheet.cell(row, col_ans, f"ERROR: {str(e)}")
            wb.save(result_path)
            torch.cuda.empty_cache()

    wb.save(result_path)
    print(f"\n鉁?娴嬭瘯瀹屾垚锛佺粨鏋滃瓨妗? {result_path}")



