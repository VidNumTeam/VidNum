import os
import shutil
import time
import torch
import numpy as np
import openpyxl
import re
import torchvision.transforms as T
from PIL import Image
from decord import VideoReader, cpu
from transformers import AutoModel, AutoTokenizer
from jinja2 import Environment, FileSystemLoader

# 瀵煎叆浣犵殑鏈湴宸ュ叿
from utils import read_xlsx_to_list, get_choice_list, is_multiple_choice_question, is_verified_correct, get_gold_answer, apply_schema_header_aliases, get_question_text
from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths

# ==========================================
# 1. 鏍稿績宸ュ叿鍑芥暟
# ==========================================

def parse_thought_and_answer(response: str):
    """瑙ｆ瀽 <think> 鍜?<answer> 鏍囩锛屽苟鍋?A-D 鍏滃簳"""
    thought, answer = "", ""
    thought_match = re.search(r'<think>(.*?)</think>', response, re.DOTALL)
    if thought_match:
        thought = thought_match.group(1).strip()
    
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.DOTALL)
    if answer_match:
        answer = answer_match.group(1).strip()
    
    # 鍏滃簳閫昏緫锛氬鏋滄病鎵惧埌鏍囩锛屾彁鍙栨渶鍚庝竴涓嚭鐜扮殑 A-D 瀛楁瘝
    if not answer:
        potential = re.findall(r'\b([A-D])\b', response)
        answer = potential[-1] if potential else response
    return thought, answer

def get_video_tensors_fast(video_path: str, input_size: int = 448):
    """48 甯ф娊甯х瓥鐣ワ細纭繚 Benchmark 姣旇緝鐨勫叕骞虫€?"""
    vr = VideoReader(video_path, ctx=cpu(0))
    fps = vr.get_avg_fps()
    duration = len(vr) / fps
    # 绛栫暐锛?=48s 鎸?1fps锛?48s 鎶?48 甯?
    num_segments = max(1, int(round(duration))) if duration <= 48.0 else 48
    frame_indices = np.linspace(0, len(vr) - 1, num_segments, dtype=int)
    
    frames_np = vr.get_batch(frame_indices).asnumpy()
    frames_tensor = torch.from_numpy(frames_np).permute(0, 3, 1, 2).float() / 255.0
    
    transform = T.Compose([
        T.Resize((input_size, input_size), interpolation=T.InterpolationMode.BICUBIC, antialias=True),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225))
    ])
    return transform(frames_tensor)


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_completed_question_set(sheet, question_col, answer_col):
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        question_en = sheet.cell(row=row, column=question_col).value
        if predicted and question_en and not str(predicted).startswith("[SKIP]") and "ERROR" not in str(predicted):
            completed.add(str(question_en).strip())
    return completed


def build_question_row_map(sheet, question_col):
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        question_en = sheet.cell(row=row, column=question_col).value
        if question_en:
            question_row_map[str(question_en).strip()] = row
    return question_row_map


def ensure_result_row(sheet, header_index, question_row_map, item):
    question_en = str(item.get("question") or "").strip()
    if question_en in question_row_map:
        return question_row_map[question_en]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Version", "Reasoning_Log", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    if question_en:
        question_row_map[question_en] = row
    return row

# ==========================================
# 2. 涓绘帹鐞嗘祦绋?
# ==========================================
if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_PATH = "OpenGVLab/InternVL3-8B"  # 鎴栨槸 InternVL3_5-8B
    MODEL_NAME = "InternVL3-8B"
    
    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    LANGUAGE = 'EN'
    TEMPLATE_PATH = "templates/qa_prompt_{}.md"

    # --- 1. 鍔犺浇妯″瀷 (閽堝鍗曞崱鐜浼樺寲) ---
    print(f"馃専 姝ｅ湪鍒濆鍖栨ā鍨? {MODEL_NAME}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH, trust_remote_code=True, use_fast=False)
    
    model = AutoModel.from_pretrained(
        MODEL_PATH,
        torch_dtype=torch.bfloat16,
        trust_remote_code=True,
        device_map="auto",
        use_flash_attn=True,   # 姝ゆ椂浣犵殑 FA2 搴旇宸茬粡瀹岀編宸ヤ綔
        low_cpu_mem_usage=False # 閬垮厤 meta tensor 鎶ラ敊
    ).eval()

    # --- 2. Excel 鍒濆鍖栦笌鏂偣缁紶 ---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    
    if os.path.exists(result_path):
        print(f"馃搨 鍙戠幇宸叉湁瀛樻。锛屾鍦ㄦ墽琛屾柇鐐圭画浼?..")
        wb = openpyxl.load_workbook(result_path)
    else:
        print(f"馃啎 鍒涘缓鏂扮粨鏋滄枃浠? {result_path}")
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        # 鍒濆鍖栬〃澶?
        base_col = sheet.max_column
        sheet.cell(1, base_col + 1, "Model_Version")
        sheet.cell(1, base_col + 2, "Reasoning_Log")
        sheet.cell(1, base_col + 3, "Predicted_Answer")
        wb.save(result_path)

    sheet = wb.active
    header_index = build_header_index(sheet)
    col_model = header_index["Model_Version"]
    col_think = header_index["Reasoning_Log"]
    col_ans = header_index["Predicted_Answer"]
    col_question_en = header_index["question"]
    completed_questions = build_completed_question_set(sheet, col_question_en, col_ans)
    question_row_map = build_question_row_map(sheet, col_question_en)

    # --- 3. 寰幆鎺ㄧ悊 ---
    gen_config = dict(max_new_tokens=1024, do_sample=False)
    jinja_env = Environment(loader=FileSystemLoader('.'))

    print(f"馃殌 寮€濮嬭窇娴嬶紝鐩爣棰樻暟: {len(data)}")

    for i, item in enumerate(data):
        question_en = str(item.get("question") or "").strip()
        if question_en and question_en in completed_questions:
            print("[LOG]")
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)

        # [閫昏緫 2]: 鍩虹杩囨护
        if not is_verified_correct(item): continue

        # [閫昏緫 3]: 鎸夐棶棰樼被鍨嬭烦杩囬潪閫夋嫨棰?
        if not is_multiple_choice_question(item):
            print(f"鈴?璺宠繃绗?{i} 棰? 妫€娴嬩负绠€绛旈 (Open-ended)")
            continue

        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        if not video_path or not os.path.exists(video_path):
            sheet.cell(row, col_ans, "[SKIP] File Missing")
            continue

        try:
            # 瑙嗚棰勫鐞?
            pixel_values = get_video_tensors_fast(video_path).to(model.device, dtype=torch.bfloat16)
            num_frames = pixel_values.shape[0]
            
            # Prompt 鍑嗗
            q_data = {
                "Question": get_question_text(item, LANGUAGE),
                "Choices": get_choice_list(item)
            }
            video_prefix = ''.join([f'Frame{j+1}: <image>\n' for j in range(num_frames)])
            prompt_text = jinja_env.get_template(TEMPLATE_PATH.format(LANGUAGE)).render(**q_data)
            
            # 妯″瀷璋冪敤
            response = model.chat(
                tokenizer, pixel_values, video_prefix + prompt_text, gen_config,
                num_patches_list=[1] * num_frames, history=None, return_history=False
            )
            if isinstance(response, tuple): response = response[0]
            
            # 瑙ｆ瀽涓庡瓨鍌?
            thought, answer = parse_thought_and_answer(response)
            
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_think, thought)
            sheet.cell(row, col_ans, answer)
            if question_en:
                completed_questions.add(question_en)
            
            print(f"鉁?[{i}/{len(data)}] 棰勬祴: {answer} | 姝ｇ‘: {get_gold_answer(item, LANGUAGE)}")
            
            # 姣?5 棰樺瓨鐩樹竴娆?
            if i % 5 == 0: wb.save(result_path)

        except Exception as e:
            print(f"鉂?绗?{i} 棰樺嚭閿? {e}")
            sheet.cell(row, col_ans, f"ERROR: {str(e)}")
            wb.save(result_path)
            torch.cuda.empty_cache()

    wb.save(result_path)
    print(f"\n馃弫 鎺ㄧ悊浠诲姟鍦嗘弧缁撴潫锛佹渶缁堝瓨妗? {result_path}")



