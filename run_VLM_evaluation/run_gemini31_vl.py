import argparse
import os
import random
import re
import threading
from zipfile import BadZipFile
from concurrent.futures import ThreadPoolExecutor, as_completed
from time import sleep, time

import openpyxl
from dotenv import load_dotenv
from google import genai
from google.genai import types

from utils import read_xlsx_to_list, render_template, get_choice_list, is_multiple_choice_question, is_verified_correct, apply_schema_header_aliases, get_question_text, get_gold_answer
from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, VIDEOS_ROOT_DEFAULT, build_paths

load_dotenv()


DEFAULT_MODELS = ["gemini-3.1-flash-preview", "gemini-3.1-pro"]
DEFAULT_TEMPLATE = "templates/qa_prompt_EN.md"
DEFAULT_WORKERS = 4
DEFAULT_SAVE_EVERY = 10
DEFAULT_MAX_RETRIES = 3
DEFAULT_RETRY_BASE_SECONDS = 1.5

_thread_local = threading.local()


def get_client() -> genai.Client:
    client = getattr(_thread_local, "client", None)
    if client is None:
        client = genai.Client(
            api_key=os.getenv("GOOGLE_API_KEY"),
            http_options={
                "base_url": os.getenv("GOOGLE_BASE_URL", "https://api.v3.cm"),
                "headers": {"Authorization": f"Bearer {os.getenv('GOOGLE_API_KEY')}"},
            },
        )
        _thread_local.client = client
    return client


def parse_answer(response: str) -> str:
    answer_match = re.search(r"<answer>(.*?)</answer>", response or "", re.IGNORECASE | re.DOTALL)
    if answer_match:
        text = answer_match.group(1).strip().upper()
        if text in {"A", "B", "C", "D"}:
            return text

    candidates = re.findall(r"\b([A-D])\b", (response or "").upper())
    if candidates:
        return candidates[-1]
    return "[ERR]"


def sanitize_model_name(model_name: str) -> str:
    safe = model_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    safe = safe.replace(" ", "_")
    return safe


def infer_prompt_mode(template_path: str) -> str:
    name = os.path.basename(template_path).lower()
    return "NoCoT" if "nocot" in name else "CoT"


def classify_error(error_text: str) -> str:
    text = (error_text or "").lower()
    if not text:
        return ""
    if "unexpected_eof_while_reading" in text or "ssl" in text:
        return "ssl_eof"
    if "server disconnected" in text:
        return "server_disconnect"
    if "10054" in text:
        return "connection_reset"
    if "429" in text or "rate limit" in text:
        return "rate_limit"
    if "403" in text or "forbidden" in text:
        return "forbidden_403"
    if "prohibited_content" in text or "prompt_blocked" in text:
        return "prompt_blocked"
    if "timed out" in text or "timeout" in text:
        return "timeout"
    return "other_error"


def is_retryable_error(error_text: str) -> bool:
    t = classify_error(error_text)
    return t in {"ssl_eof", "server_disconnect", "connection_reset", "timeout", "rate_limit"}


def resolve_video_path(item: dict, datacuts_root: str, videos_root: str) -> str:
    _, _, full_video_path, cut_video_path = build_paths(
        item,
        videos_root=videos_root,
        datacuts_root=datacuts_root,
    )
    if os.path.exists(cut_video_path):
        return cut_video_path
    return full_video_path


def build_item_key(item: dict) -> str:
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def ensure_result_headers(sheet):
    header_index = build_header_index(sheet)
    required = [
        "Model_Version",
        "Full_Response",
        "Predicted_Answer",
        "Status",
        "Error",
        "Error_Type",
        "Latency_Seconds",
        "Video_Path",
    ]
    if all(h in header_index for h in required):
        return header_index

    next_col = sheet.max_column + 1
    for h in required:
        if h not in header_index:
            sheet.cell(row=1, column=next_col, value=h)
            next_col += 1
    return build_header_index(sheet)


def build_row_map(sheet, header_index):
    row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value if "ID" in header_index else None,
            "question": sheet.cell(row=row, column=header_index["question"]).value if "question" in header_index else None,
        }
        row_map[build_item_key(row_item)] = row
    return row_map


def ensure_row(sheet, header_index, row_map, item):
    key = build_item_key(item)
    if key in row_map:
        return row_map[key]

    row = sheet.max_row + 1
    for h, c in header_index.items():
        if h in {
            "Model_Version",
            "Full_Response",
            "Predicted_Answer",
            "Status",
            "Error",
            "Error_Type",
            "Latency_Seconds",
            "Video_Path",
        }:
            continue
        sheet.cell(row=row, column=c, value=item.get(h))
    row_map[key] = row
    return row


def build_done_keys(sheet, header_index):
    done = set()
    col_status = header_index["Status"]
    col_pred = header_index["Predicted_Answer"]
    for row in range(2, sheet.max_row + 1):
        status = str(sheet.cell(row=row, column=col_status).value or "").strip().lower()
        pred = str(sheet.cell(row=row, column=col_pred).value or "").strip()
        if status == "ok" and pred not in ("", "[ERR]"):
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value if "ID" in header_index else None,
                "question": sheet.cell(row=row, column=header_index["question"]).value if "question" in header_index else None,
            }
            done.add(build_item_key(row_item))
    return done


def infer_one(
    item: dict,
    model_name: str,
    template_path: str,
    datacuts_root: str,
    videos_root: str,
    max_retries: int,
    retry_base_seconds: float,
):
    key = build_item_key(item)
    video_path = resolve_video_path(item, datacuts_root=datacuts_root, videos_root=videos_root)
    if not os.path.exists(video_path):
        return key, {
            "full_response": "",
            "answer": "",
            "status": "video_missing",
            "error": "video file not found",
            "latency": 0.0,
            "video_path": video_path,
        }

    q_data = {
        "Question": get_question_text(item, LANGUAGE if "LANGUAGE" in globals() else "EN"),
        "Choices": get_choice_list(item),
    }
    prompt_text = render_template(template_path, q_data)

    with open(video_path, "rb") as f:
        video_bytes = f.read()
    video_part = types.Part.from_bytes(data=video_bytes, mime_type="video/mp4")

    started = time()
    client = get_client()
    last_err = ""
    for attempt in range(max_retries + 1):
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=[video_part, prompt_text],
                config=types.GenerateContentConfig(temperature=0.0),
            )
            full = response.text or ""
            ans = parse_answer(full)
            latency = round(time() - started, 3)
            status = "ok" if ans != "[ERR]" else "bad_format"
            return key, {
                "full_response": full,
                "answer": ans,
                "status": status,
                "error": "",
                "latency": latency,
                "video_path": video_path,
            }
        except Exception as exc:
            last_err = str(exc)
            if attempt >= max_retries or not is_retryable_error(last_err):
                break
            backoff = retry_base_seconds * (2**attempt) + random.uniform(0, 0.4)
            sleep(backoff)

    return key, {
        "full_response": "",
        "answer": "",
        "status": "error",
        "error": last_err,
        "latency": round(time() - started, 3),
        "video_path": video_path,
    }


def run_for_model(
    data: list[dict],
    model_name: str,
    template_path: str,
    workers: int,
    data_path: str,
    datacuts_root: str,
    videos_root: str,
    save_every: int,
    max_retries: int,
    retry_base_seconds: float,
):
    safe_model = sanitize_model_name(model_name)
    mode_tag = infer_prompt_mode(template_path)
    result_path = data_path.replace(".xlsx", f"_{safe_model}_{mode_tag}_Results_gsdk.xlsx")

    if os.path.exists(result_path):
        try:
            wb = openpyxl.load_workbook(result_path)
        except BadZipFile:
            backup_path = result_path + ".corrupt"
            os.replace(result_path, backup_path)
            print(f"[WARN] Corrupted result file detected. Moved to: {backup_path}")
            wb = openpyxl.load_workbook(data_path)
    else:
        wb = openpyxl.load_workbook(data_path)
    sheet = wb.active
    header_index = ensure_result_headers(sheet)
    row_map = build_row_map(sheet, header_index)
    done_keys = build_done_keys(sheet, header_index)
    wb.save(result_path)

    pending = []
    for item in data:
        if not is_multiple_choice_question(item):
            continue
        if not is_verified_correct(item):
            continue
        key = build_item_key(item)
        if key in done_keys:
            continue
        pending.append(item)

    print(f"\n=== {model_name} ===")
    print(f"Prompt mode: {mode_tag} (template={template_path})")
    print(f"Result file: {result_path}")
    print(f"Pending: {len(pending)} | workers={workers}")
    if not pending:
        wb.close()
        return

    col_model = header_index["Model_Version"]
    col_full = header_index["Full_Response"]
    col_pred = header_index["Predicted_Answer"]
    col_status = header_index["Status"]
    col_err = header_index["Error"]
    col_err_type = header_index["Error_Type"]
    col_lat = header_index["Latency_Seconds"]
    col_video = header_index["Video_Path"]

    processed = 0
    evaluated = 0
    correct = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(
                infer_one,
                item=item,
                model_name=model_name,
                template_path=template_path,
                datacuts_root=datacuts_root,
                videos_root=videos_root,
                max_retries=max_retries,
                retry_base_seconds=retry_base_seconds,
            ): item
            for item in pending
        }
        for future in as_completed(futures):
            item = futures[future]
            try:
                key, out = future.result()
            except Exception as exc:
                key = build_item_key(item)
                out = {
                    "full_response": "",
                    "answer": "",
                    "status": "error",
                    "error": str(exc),
                    "latency": 0.0,
                    "video_path": "",
                }

            row = ensure_row(sheet, header_index, row_map, item)
            sheet.cell(row=row, column=col_model, value=model_name)
            sheet.cell(row=row, column=col_full, value=out["full_response"])
            sheet.cell(row=row, column=col_pred, value=out["answer"])
            sheet.cell(row=row, column=col_status, value=out["status"])
            sheet.cell(row=row, column=col_err, value=out["error"])
            sheet.cell(row=row, column=col_err_type, value=classify_error(str(out["error"])))
            sheet.cell(row=row, column=col_lat, value=out["latency"])
            sheet.cell(row=row, column=col_video, value=out["video_path"])
            done_keys.add(key)

            processed += 1
            pred = str(out["answer"] or "").strip().upper()
            gold = str(item.get("绛旀_EN") or item.get("绛旀") or "").strip().upper()
            if pred in {"A", "B", "C", "D"} and gold in {"A", "B", "C", "D"}:
                evaluated += 1
                if pred == gold:
                    correct += 1

            if processed % save_every == 0:
                wb.save(result_path)
            acc = (correct / evaluated * 100.0) if evaluated else 0.0
            err_preview = str(out.get("error") or "").strip().replace("\n", " ")
            if len(err_preview) > 120:
                err_preview = err_preview[:117] + "..."
            print(
                f"[{processed}/{len(pending)}] ID={item.get('ID')} "
                f"ans={out['answer']} gold={gold} status={out['status']} "
                f"| acc={correct}/{evaluated} ({acc:.2f}%)"
                + (f" | err={err_preview}" if err_preview else "")
            )

    wb.save(result_path)
    wb.close()
    final_acc = (correct / evaluated * 100.0) if evaluated else 0.0
    print(
        f"Completed {model_name}. Saved: {result_path} | "
        f"final acc={correct}/{evaluated} ({final_acc:.2f}%)"
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    parser.add_argument("--template", type=str, default=DEFAULT_TEMPLATE)
    parser.add_argument("--data", type=str, default=DATA_PATH_DEFAULT)
    parser.add_argument("--models", type=str, default=",".join(DEFAULT_MODELS), help="comma separated model ids")
    parser.add_argument("--save-every", type=int, default=DEFAULT_SAVE_EVERY)
    parser.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES)
    parser.add_argument("--retry-base-seconds", type=float, default=DEFAULT_RETRY_BASE_SECONDS)
    parser.add_argument("--datacuts-root", type=str, default=DATACUTS_ROOT_DEFAULT)
    parser.add_argument("--videos-root", type=str, default=VIDEOS_ROOT_DEFAULT)
    args = parser.parse_args()

    if args.workers < 1:
        raise ValueError("--workers must be >= 1")
    if args.save_every < 1:
        raise ValueError("--save-every must be >= 1")
    if args.max_retries < 0:
        raise ValueError("--max-retries must be >= 0")
    if args.retry_base_seconds < 0:
        raise ValueError("--retry-base-seconds must be >= 0")

    models = [m.strip() for m in args.models.split(",") if m.strip()]
    if not models:
        raise ValueError("No models provided")

    data = read_xlsx_to_list(args.data)
    print(f"Loaded {len(data)} rows from {args.data}")
    for model_name in models:
        run_for_model(
            data=data,
            model_name=model_name,
            template_path=args.template,
            workers=args.workers,
            data_path=args.data,
            datacuts_root=args.datacuts_root,
            videos_root=args.videos_root,
            save_every=args.save_every,
            max_retries=args.max_retries,
            retry_base_seconds=args.retry_base_seconds,
        )


if __name__ == "__main__":
    main()



