import os
import torch
import openpyxl
import re
import shutil
import av
import numpy as np
import sys
from transformers import AutoProcessor, LlavaNextVideoForConditionalGeneration
from jinja2 import Environment, FileSystemLoader
from utils import get_choice_list, is_multiple_choice_question, is_verified_correct, get_gold_answer, apply_schema_header_aliases, get_question_text

# ==========================================
# 1. 瑙嗛瑙ｇ爜宸ュ叿 (PyAV)
# ==========================================
def read_video_pyav(container, indices):
    frames = []
    container.seek(0)
    start_index = indices[0]
    end_index = indices[-1]
    for i, frame in enumerate(container.decode(video=0)):
        if i > end_index:
            break
        if i >= start_index and i in indices:
            frames.append(frame)
    # 纭繚杩斿洖鐨勫抚鏁颁笌绱㈠紩鏁颁竴鑷达紝澶勭悊涓埆瑙嗛瑙ｇ爜璺冲抚闂
    if len(frames) == 0:
        return None
    return np.stack([x.to_ndarray(format="rgb24") for x in frames])

# ==========================================
# 2. 瑙ｆ瀽閫昏緫 (鎻愬彇 Thinking 鍜?Answer)
# ==========================================
def parse_answer(response: str):
    """鍚屾椂鎻愬彇鎬濈淮閾?think)鍜屾渶缁堢瓟妗?answer)"""
    # 1. 鎻愬彇 <think> 鏍囩鍐呭
    thought_match = re.search(r'<think>(.*?)</think>', response, re.IGNORECASE | re.DOTALL)
    thought_process = thought_match.group(1).strip() if thought_match else ""

    # 2. 鎻愬彇 <answer> 鏍囩鍐呭
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.IGNORECASE | re.DOTALL)
    if answer_match:
        final_answer = answer_match.group(1).strip().upper()
    else:
        # 鍏滃簳锛氬鎵惧洖澶嶆湯灏惧嚭鐜扮殑 A-D
        potential = re.findall(r'\b([A-D])\b', response)
        final_answer = potential[-1].upper() if potential else "[ERR]"
    
    return thought_process, final_answer


# ==========================================
# 2. 椴佹瑙ｆ瀽閫昏緫锛氭敮鎸佸绉?CoT 鏍煎紡
# ==========================================
def parse_answer(response: str):
    """
    瑙ｆ瀽鎺ㄧ悊杩囩▼鍜屾渶缁堢瓟妗?
    鏀寔 鏍囩銆?## Analysis 鏍囬鎴栫函鏂囨湰鍏滃簳
    """
    thought, answer = "", "[ERR]"
    

    if "Answer" in response:
        parts = response.split("### Answer:")
        thought = parts[0].strip()
        if len(parts) > 1:
            ans_text = parts[1].strip().upper()
            match = re.search(r'\b([A-D])\b', ans_text)
            if match: answer = match.group(1)

    # C. 鍏滃簳锛氬鎵炬渶鍚庝竴涓嚭鐜扮殑 A/B/C/D
    if answer == "[ERR]":
        potential = re.findall(r'\b([A-D])\b', response.upper())
        if potential:
            answer = potential[-1]
            thought = response.replace(answer, "").strip()

    return thought, answer


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_item_key(item):
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"
def build_completed_question_set(sheet, header_index, answer_col):
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        predicted_text = str(predicted).strip() if predicted is not None else ""
        if predicted_text and not predicted_text.startswith("[SKIP]") and "ERROR" not in predicted_text and predicted_text != "[ERR]":
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value,
                "question": sheet.cell(row=row, column=header_index["question"]).value,
            }
            completed.add(build_item_key(row_item))
    return completed


def build_question_row_map(sheet, header_index):
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value,
            "question": sheet.cell(row=row, column=header_index["question"]).value,
        }
        question_row_map[build_item_key(row_item)] = row
    return question_row_map


def ensure_result_row(sheet, header_index, question_row_map, item):
    item_key = build_item_key(item)
    if item_key in question_row_map:
        return question_row_map[item_key]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Version", "Model_Thinking", "Full_Response", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    question_row_map[item_key] = row
    return row


# ==========================================
# 3. 涓绘帹鐞嗘祦绋?
# ==========================================
if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_ID = "llava-hf/LLaVA-NeXT-Video-7B-hf"
    MODEL_NAME = "LLaVA-NeXT-7B-NoCoT"
    
    try:
        from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths
        from utils import read_xlsx_to_list
    except ImportError:
        print("[LOG]")
        sys.exit(1)

    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    TEMPLATE_PATH = "templates/qa_prompt_llava_noCoT.md" 

    # --- 1. 鍔犺浇妯″瀷 (FP16 浼樺寲) ---
    print(f"馃殌 姝ｅ湪鍔犺浇 {MODEL_NAME}...")
    model = LlavaNextVideoForConditionalGeneration.from_pretrained(
        MODEL_ID, 
        torch_dtype=torch.float16, 
        low_cpu_mem_usage=True, 
        device_map="auto"
    )
    processor = AutoProcessor.from_pretrained(MODEL_ID)

    # --- 2. Excel 鏂偣缁紶鍑嗗 ---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    
    if os.path.exists(result_path):
        print(f"馃搨 鍙戠幇宸叉湁缁撴灉锛屾鍦ㄦ柇鐐圭画浼?..")
        wb = openpyxl.load_workbook(result_path)
    else:
        print(f"馃啎 鍒涘缓鏂扮粨鏋滄枃浠?..")
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        base_col = sheet.max_column
        headers = ["Model_Version", "Model_Thinking", "Full_Response", "Predicted_Answer"]
        for idx, h in enumerate(headers):
            sheet.cell(1, base_col + idx + 1, h)
        wb.save(result_path)

    sheet = wb.active
    header_index = build_header_index(sheet)
    col_model = header_index["Model_Version"]
    col_think = header_index["Model_Thinking"]
    col_full = header_index["Full_Response"]
    col_ans = header_index["Predicted_Answer"]
    completed_questions = build_completed_question_set(sheet, header_index, col_ans)
    question_row_map = build_question_row_map(sheet, header_index)

    jinja_env = Environment(loader=FileSystemLoader('.'))

    # --- 3. 寰幆娴嬭瘯 ---
    for i, item in enumerate(data):
        item_key = build_item_key(item)
        if item_key in completed_questions:
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)

        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        abs_video_path = os.path.abspath(video_path)
        if not os.path.exists(abs_video_path): continue

        try:
            # 娓叉煋鎻愮ず璇?
            q_data = {
                "Question": get_question_text(item, LANGUAGE if "LANGUAGE" in globals() else "EN"),
                "Choices": get_choice_list(item)
            }
            prompt_raw = jinja_env.get_template(TEMPLATE_PATH).render(**q_data)
            prompt = f"USER: <video>\n{prompt_raw}"

            # --- [瀵归綈 InternVL 鐨勬娊甯ч€昏緫] ---
            container = av.open(abs_video_path)
            # 鑾峰彇瑙嗛娴?
            video_stream = container.streams.video[0]
            total_frames = video_stream.frames
            # 璁＄畻鏃堕暱
            avg_fps = float(video_stream.average_rate)
            duration = total_frames / avg_fps if avg_fps > 0 else 0
            
            # 閫昏緫锛? FPS 閲囨牱锛屼笂闄?48 甯?
            if duration <= 48:
                num_frames_to_sample = max(1, int(duration))
            else:
                num_frames_to_sample = 48
            
            # 鐢熸垚绛夐棿璺濈储寮?
            indices = np.linspace(0, total_frames - 1, num_frames_to_sample).astype(int)
            clip = read_video_pyav(container, indices)
            container.close()

            if clip is None:
                print(f"鉂?绱㈠紩 {i}: 瑙嗛瑙ｇ爜澶辫触")
                continue

            # 棰勫鐞?
            inputs = processor(text=prompt, videos=clip, return_tensors="pt").to(model.device)

            # --- [鎺ㄧ悊鐢熸垚] ---
            with torch.no_grad():
                generate_ids = model.generate(
                    **inputs, 
                    max_new_tokens=512, # 闄愬埗鐢熸垚闀垮害
                    do_sample=False
                )
            
            # 瑙ｇ爜涓庢竻娲?
            full_response = processor.batch_decode(generate_ids, skip_special_tokens=True, clean_up_tokenization_spaces=False)[0]
            
            # 绉婚櫎 Prompt 閮ㄥ垎
            if "ASSISTANT:" in full_response:
                clean_response = full_response.split("ASSISTANT:")[-1].strip()
            else:
                clean_response = full_response.strip()

            # 瑙ｆ瀽缁撴灉
            thought, answer = parse_answer(clean_response)

            # 鍐欏叆缁撴灉
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_think, thought)
            sheet.cell(row, col_full, clean_response)
            sheet.cell(row, col_ans, answer)
            completed_questions.add(item_key)

            print(f"鉁?[{i+1}/{len(data)}] 棰勬祴: {answer} (閲囨牱 {num_frames_to_sample} 甯?")
            
            if (i + 1) % 5 == 0: wb.save(result_path)

        except Exception as e:
            print(f"鉂?绱㈠紩 {i} 鎺ㄧ悊寮傚父: {e}")
            import traceback
            traceback.print_exc()
            torch.cuda.empty_cache()

    wb.save(result_path)
    print("[LOG]")



