import os
import torch
import openpyxl
import re
import shutil
from transformers import Qwen2_5_VLForConditionalGeneration, AutoProcessor
from qwen_vl_utils import process_vision_info
from jinja2 import Environment, FileSystemLoader
from utils import get_choice_list, is_multiple_choice_question, is_verified_correct, get_gold_answer, apply_schema_header_aliases, get_question_text

# 瀵煎叆浣犵殑鏈湴宸ュ叿
try:
    from utils import read_xlsx_to_list
    from videocut_multithread import DATACUTS_ROOT_DEFAULT, DATA_PATH_DEFAULT, build_paths
except ImportError:
    print("[LOG]")

# ==========================================
# 1. 绛旀瑙ｆ瀽閫昏緫
# ==========================================
def parse_answer(response: str):
    """鎻愬彇鏈€缁堢殑 A/B/C/D 绛旀"""
    # 浼樺厛鍖归厤 <answer> 鏍囩
    answer_match = re.search(r'<answer>(.*?)</answer>', response, re.IGNORECASE | re.DOTALL)
    if answer_match:
        return answer_match.group(1).strip().upper()
    
    # 鍏滃簳锛氬鎵惧洖澶嶄腑鏈€鍚庡嚭鐜扮殑 A-D 瀛楁瘝
    potential = re.findall(r'\b([A-D])\b', response)
    if potential:
        return potential[-1].upper()
    
    return response.strip()[:10]


def build_header_index(sheet):
    header_index = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_index[str(value).strip()] = col
    return apply_schema_header_aliases(header_index)


def build_item_key(item):
    item_id = item.get("ID")
    if item_id not in (None, ""):
        return f"ID::{item_id}"
    question_text = str(item.get("question") or "").strip()
    return f"Q::{question_text}"
def build_completed_question_set(sheet, header_index, answer_col):
    completed = set()
    for row in range(2, sheet.max_row + 1):
        predicted = sheet.cell(row=row, column=answer_col).value
        predicted_text = str(predicted).strip() if predicted is not None else ""
        if predicted_text and predicted_text != "[ERR]":
            row_item = {
                "ID": sheet.cell(row=row, column=header_index["ID"]).value,
                "question": sheet.cell(row=row, column=header_index["question"]).value,
            }
            completed.add(build_item_key(row_item))
    return completed


def build_question_row_map(sheet, header_index):
    question_row_map = {}
    for row in range(2, sheet.max_row + 1):
        row_item = {
            "ID": sheet.cell(row=row, column=header_index["ID"]).value,
            "question": sheet.cell(row=row, column=header_index["question"]).value,
        }
        question_row_map[build_item_key(row_item)] = row
    return question_row_map


def ensure_result_row(sheet, header_index, question_row_map, item):
    item_key = build_item_key(item)
    if item_key in question_row_map:
        return question_row_map[item_key]

    row = sheet.max_row + 1
    for header, col in header_index.items():
        if header in {"Model_Version", "Full_Response", "Predicted_Answer"}:
            continue
        sheet.cell(row=row, column=col, value=item.get(header))
    question_row_map[item_key] = row
    return row

# ==========================================
# 2. 涓绘帹鐞嗘祦绋?
# ==========================================
if __name__ == "__main__":
    # --- [閰嶇疆鍖篯 ---
    MODEL_ID = "Qwen/Qwen2.5-VL-7B-Instruct"
    MODEL_NAME = "Qwen2.5-VL-7B-NoCoT"
    
    DATA_PATH = DATA_PATH_DEFAULT
    DATACUTS_ROOT = DATACUTS_ROOT_DEFAULT
    LANGUAGE = 'EN'
    TEMPLATE_PATH = "templates/qa_prompt_EN_noCoT.md"

    # --- 1. 鍔犺浇妯″瀷 (鍙?4090 浼樺寲) ---
    print(f"馃専 姝ｅ湪鍔犺浇 Qwen2.5-VL-7B... 鐩爣锛氬弻 4090 鑷姩骞惰")
    
    # Qwen2.5-VL 鍘熺敓鏀寔 flash_attention_2
    model = Qwen2_5_VLForConditionalGeneration.from_pretrained(
        MODEL_ID,
        torch_dtype=torch.bfloat16,
        attn_implementation="flash_attention_2",
        device_map="auto",  # 鑷姩鍒囧垎鍒颁袱寮?4090
    )
    
    # 璁惧畾瑙嗛澶勭悊鐨勫儚绱犳瀬闄愶紝闃叉闀胯棰?OOM
    # Qwen2.5-VL 鏀寔鍔ㄦ€佸垎杈ㄧ巼锛岃繖閲岃缃竴涓悎鐞嗙殑涓婇檺
    processor = AutoProcessor.from_pretrained(MODEL_ID)

    # --- 2. 鍑嗗 Excel 鏂偣缁紶 ---
    data = read_xlsx_to_list(DATA_PATH)
    result_path = DATA_PATH.replace(".xlsx", f"_{MODEL_NAME}_Results_fixed.xlsx")
    
    if os.path.exists(result_path):
        print(f"馃搨 鍙戠幇宸叉湁缁撴灉锛屾柇鐐圭画浼犱腑...")
        wb = openpyxl.load_workbook(result_path)
    else:
        print(f"馃啎 鍒涘缓鏂扮粨鏋滄枃浠?..")
        shutil.copy2(DATA_PATH, result_path)
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        # 鍒濆鍖栬〃澶?
        base_col = sheet.max_column
        sheet.cell(1, base_col + 1, "Model_Version")
        sheet.cell(1, base_col + 2, "Full_Response")
        sheet.cell(1, base_col + 3, "Predicted_Answer")
        wb.save(result_path)

    sheet = wb.active
    header_index = build_header_index(sheet)
    col_model = header_index["Model_Version"]
    col_full = header_index["Full_Response"]
    col_ans = header_index["Predicted_Answer"]
    completed_questions = build_completed_question_set(sheet, header_index, col_ans)
    question_row_map = build_question_row_map(sheet, header_index)

    jinja_env = Environment(loader=FileSystemLoader('.'))

    print(f"馃殌 寮€濮嬫祴璇?..")

    for i, item in enumerate(data):
        # 杩囨护閫昏緫
        if not is_multiple_choice_question(item): 
            print(f"鈿狅笍 璺宠繃闈為€夋嫨棰? {str(get_question_text(item, LANGUAGE if 'LANGUAGE' in globals() else 'EN'))[:30]}...")
            continue

        if not is_verified_correct(item): continue
        item_key = build_item_key(item)
        if item_key in completed_questions:
            print(f"鉁?宸插畬鎴愶紝璺宠繃: {str(get_question_text(item, LANGUAGE if 'LANGUAGE' in globals() else 'EN'))[:30]}...")
            continue

        row = ensure_result_row(sheet, header_index, question_row_map, item)

        video_path = build_paths(item, datacuts_root=DATACUTS_ROOT)[3]
        if not video_path or not os.path.exists(video_path):
            print(f"鈿狅笍 鎵句笉鍒拌棰? {video_path}")
            continue

        try:
            # 鍑嗗鍐呭
            q_data = {
                "Question": get_question_text(item, LANGUAGE),
                "Choices": get_choice_list(item)
            }
            prompt_text = jinja_env.get_template(TEMPLATE_PATH).render(**q_data)

            # Qwen2.5-VL 鏍囧噯杈撳叆鏍煎紡
            messages = [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "video",
                            "video": os.path.abspath(video_path),
                            "fps": 1.0, # 鎶藉抚棰戠巼锛屾牴鎹棰戦暱搴﹁皟鏁?
                        },
                        {"type": "text", "text": prompt_text},
                    ],
                }
            ]

            # 棰勫鐞?
            text = processor.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
            image_inputs, video_inputs = process_vision_info(messages)
            
            inputs = processor(
                text=[text],
                images=image_inputs,
                videos=video_inputs,
                padding=True,
                return_tensors="pt",
            )
            inputs = inputs.to(model.device)

            # 鎺ㄧ悊
            with torch.no_grad():
                generated_ids = model.generate(**inputs, max_new_tokens=512)
            
            # 鍓旈櫎 Prompt 閮ㄥ垎
            generated_ids_trimmed = [
                out_ids[len(in_ids) :] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
            ]
            response = processor.batch_decode(
                generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
            )[0]

            answer = parse_answer(response)
            
            # 鍐欏叆骞朵繚瀛?
            sheet.cell(row, col_model, MODEL_NAME)
            sheet.cell(row, col_full, response)
            sheet.cell(row, col_ans, answer)
            completed_questions.add(item_key)
            
            print(f"鉁?[{i+1}/{len(data)}] 棰勬祴: {answer}")
            
            # if (i + 1) % 5 == 0:
            wb.save(result_path)

        except Exception as e:
            print(f"鉂?绱㈠紩 {i} 鎶ラ敊: {e}")
            torch.cuda.empty_cache()

    wb.save(result_path)
    print("[LOG]")



