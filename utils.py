from jinja2 import Environment, FileSystemLoader
import os
import openpyxl
import openai
from dotenv import load_dotenv

load_dotenv()


def _copy_alias_if_missing(item: dict, target_key: str, candidate_keys: list[str]):
    if target_key in item and item.get(target_key) not in (None, ""):
        return
    for key in candidate_keys:
        if key in item and item.get(key) not in (None, ""):
            item[target_key] = item.get(key)
            return


def add_schema_aliases(item: dict) -> dict:
    """New-schema adapter for run scripts."""
    _copy_alias_if_missing(item, "Question_EN", ["question"])
    _copy_alias_if_missing(item, "Question", ["question"])
    _copy_alias_if_missing(item, "ChoiceA", ["option_A"])
    _copy_alias_if_missing(item, "ChoiceB", ["option_B"])
    _copy_alias_if_missing(item, "ChoiceC", ["option_C"])
    _copy_alias_if_missing(item, "ChoiceD", ["option_D"])
    _copy_alias_if_missing(item, "ChoiceA.1", ["option_B"])
    _copy_alias_if_missing(item, "ChoiceA.2", ["option_C"])
    _copy_alias_if_missing(item, "ChoiceA.3", ["option_D"])

    if item.get("Video_Path") in (None, ""):
        item_id = item.get("ID")
        if item_id not in (None, ""):
            item["Video_Path"] = f"QID_{item_id}.mp4"

    return item


def apply_schema_header_aliases(header_index: dict) -> dict:
    """New-schema adapter for result headers."""
    alias_groups = {
        "Question_EN": ["question"],
        "Question": ["question"],
        "ChoiceA": ["option_A"],
        "ChoiceB": ["option_B"],
        "ChoiceC": ["option_C"],
        "ChoiceD": ["option_D"],
        "ChoiceA.1": ["option_B"],
        "ChoiceA.2": ["option_C"],
        "ChoiceA.3": ["option_D"],
    }
    for canonical, candidates in alias_groups.items():
        if canonical in header_index:
            continue
        for candidate in candidates:
            if candidate in header_index:
                header_index[canonical] = header_index[candidate]
                break
    return header_index

def get_question_text(item: dict, language: str = "EN") -> str:
    _ = language
    return str(item.get("question") or "").strip()



def LLM_response(prompt: str, model: str = "gemini-3") -> str:
    """
    璋冪敤澶ц瑷€妯″瀷鑾峰彇鍥炲銆?
    
    :param prompt: 鎻愮ず璇嶅唴瀹?
    :param model: 妯″瀷鍚嶇О
    :return: 妯″瀷鐢熸垚鐨勬枃鏈唴瀹?
    """
    try:
        response = openai.chat.completions.create(
            model=model,
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0,    
        )
        
        answer_text = response.choices[0].message.content
        return answer_text
        
    except Exception as e:
        # 娣诲姞鍩虹鐨勯敊璇鐞嗭紝闃叉鏌愪竴鏉¤姹傚け璐ュ鑷存暣涓惊鐜腑鏂?
        print(f"璋冪敤妯″瀷鏃跺彂鐢熼敊璇? {e}")
        return f"Error: {e}"

def render_template(template_path: str, context: dict) -> str:
    """
    浣跨敤 Jinja2 鍔犺浇骞舵覆鏌撴ā鏉?(鏀寔鐩存帴浼犲叆甯﹁矾寰勭殑妯℃澘鏂囦欢)銆?
    
    :param template_path: 妯℃澘鏂囦欢鐨勭浉瀵硅矾寰勬垨缁濆璺緞 (渚嬪 'templates/property.md')
    :param context: 鐢ㄤ簬娓叉煋妯℃澘鐨勬暟鎹瓧鍏?
    :return: 娓叉煋鍚庣殑瀛楃涓?
    """
    # 鑷姩鍒嗙 "鐩綍" 鍜?"鏂囦欢鍚?
    # 渚嬪锛?templates\property.md" 浼氳鎷嗗垎涓?-> dir: "templates", name: "property.md"
    template_dir, template_name = os.path.split(template_path)
    
    # 濡傛灉璺緞涓病鏈夌洰褰曪紙姣斿鐩存帴浼犱簡 "property.md"锛夛紝鍒欓粯璁ゅ湪褰撳墠鐩綍('.')涓嬪鎵?
    if not template_dir:
        template_dir = '.'
        
    # 鍒涘缓 Jinja2 鐜锛屽姞杞藉櫒鎸囧悜鎻愬彇鍑虹殑鐩綍
    env = Environment(loader=FileSystemLoader(template_dir))
    
    # 鑾峰彇妯℃澘鏂囦欢
    template = env.get_template(template_name)
    
    # 娓叉煋骞惰繑鍥炵粨鏋?
    return template.render(**context)



def read_xlsx_to_list(file_path: str, sheet_name: str = None) -> list[dict]:
    """
    璇诲彇 xlsx 鏂囦欢锛屾寜琛岃В鏋愪负瀛楀吀鍒楄〃銆傞粯璁ょ涓€琛屼负琛ㄥご(Keys)銆?
    
    :param file_path: Excel 鏂囦欢鐨勮矾寰?
    :param sheet_name: 鎸囧畾瑕佽鍙栫殑 Sheet 鍚嶇О銆傚鏋滀笉浼狅紝鍒欒鍙栧綋鍓嶆縺娲荤殑 Sheet
    :return: 鍖呭惈姣忎竴琛屾暟鎹殑瀛楀吀鍒楄〃
    """
    # data_only=True 纭繚璇诲彇鐨勬槸鍏紡璁＄畻鍚庣殑鍊硷紝鑰屼笉鏄叕寮忔湰韬?
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active # 榛樿鑾峰彇婵€娲荤殑/绗竴涓?sheet

    data_list = []
    headers = []
    
    # 鎸夎杩唬鑾峰彇鍊?
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            # 绗竴琛屼綔涓哄瓧鍏哥殑閿?(澶勭悊琛ㄥご鍙兘鍖呭惈 None 鐨勬儏鍐?
            headers = [str(cell).strip() if cell else f"Column_{j}" for j, cell in enumerate(row)]
        else:
            # 灏嗗綋鍓嶈鏁版嵁涓庤〃澶存墦鍖呮垚瀛楀吀
            row_dict = {headers[j]: cell for j, cell in enumerate(row)}
            row_dict = add_schema_aliases(row_dict)
            
            # (鍙€? 濡傛灉浣犱笉鎯虫妸鍏ㄤ负绌虹殑绌鸿鍔犺繘鍘伙紝鍙互鍔犱釜鍒ゆ柇
            if any(value is not None for value in row_dict.values()):
                data_list.append(row_dict)
                
    # 璁板緱鍏抽棴鏂囦欢閲婃斁鍐呭瓨
    wb.close()
            
    return data_list


def pick_first(item: dict, keys: list[str], default=None):
    """Return first non-empty value from candidate keys."""
    for k in keys:
        if k in item and item.get(k) not in (None, ""):
            return item.get(k)
    return default


def get_choice_list(item: dict) -> list:
    """New naming only: option_A/B/C/D."""
    a = item.get("option_A")
    b = item.get("option_B")
    c = item.get("option_C")
    d = item.get("option_D")
    return [a, b, c, d]
def is_multiple_choice_question(item: dict) -> bool:
    """Compatible question-type filter across Chinese/English schemas."""
    val = pick_first(item, ["闂绫诲瀷", "QuestionType", "Question Type", "Type", "棰樺瀷"], default=None)
    if val is None:
        return True
    s = str(val).strip().lower()
    if s in {"single_choice", "single choice", "multiple_choice", "multiple choice", "mcq"}:
        return True
    return "閫夋嫨" in str(val)


def is_verified_correct(item: dict) -> bool:
    """Compatible correctness filter across Chinese/English schemas."""
    val = pick_first(item, ["鏍稿缁撴灉", "Verification", "Verified", "CheckResult"], default=None)
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in {"姝ｇ‘", "correct", "true", "1", "yes", "y"}


def get_gold_answer(item: dict, language: str = "EN"):
    """Get ground-truth answer with schema fallback."""
    if str(language).upper() == "EN":
        return pick_first(item, ["answer", "Answer", "Answer_EN", "绛旀_EN", "绛旀"])
    return pick_first(item, ["answer", "Answer", "绛旀", "Answer_EN", "绛旀_EN"])

def write_list_to_xlsx(data: list[dict], output_path: str):
    """
    灏嗗瓧鍏稿垪琛ㄥ啓鍥炲埌 xlsx 鏂囦欢涓€?
    
    :param data: 鍖呭惈瀛楀吀鐨勫垪琛?
    :param output_path: 淇濆瓨鐨?Excel 鏂囦欢璺緞
    """
    if not data:
        print("[LOG]")
        return
        
    # 鍒涘缓涓€涓柊鐨勫伐浣滅翱
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # 鑾峰彇琛ㄥご (鍗冲瓧鍏哥殑鎵€鏈夐敭)
    # 鍥犱负浣犲湪寰幆閲岀粰 item 鍔犱簡 'ID'锛屾墍浠?'ID' 涔熶細鍖呭惈鍦ㄨ繖浜涢敭閲岋紝閫氬父鍦ㄦ渶鍚庝竴鍒?
    headers = list(data[0].keys())
    
    # 鍐欏叆绗竴琛岋細琛ㄥご
    sheet.append(headers)
    
    # 閬嶅巻鏁版嵁锛屾寜琛ㄥご椤哄簭鎻愬彇鍊煎苟鍐欏叆姣忎竴琛?
    for item in data:
        # 浣跨敤 item.get(h) 闃叉鏌愪竴琛岀己澶辨煇涓敭鑰屾姤閿?
        row = [item.get(h) for h in headers]
        sheet.append(row)
        
    # 淇濆瓨鏂囦欢
    wb.save(output_path)



def append_row_to_xlsx(file_path: str, item: dict):
    """
    灏嗗崟鏉″瓧鍏告暟鎹拷鍔犲埌 Excel 鏂囦欢鏈熬銆?
    濡傛灉鏂囦欢涓嶅瓨鍦紝浼氳嚜鍔ㄥ垱寤哄苟浠ュ瓧鍏哥殑閿綔涓鸿〃澶淬€?
    濡傛灉鏂囦欢宸插瓨鍦紝浼氳鍙栧師鏈夎〃澶村苟瀵归綈鍐欏叆銆?
    
    :param file_path: 瑕佷繚瀛樼殑 Excel 鏂囦欢璺緞
    :param item: 鍗曟潯鏁版嵁锛堝瓧鍏革級
    """
    if os.path.exists(file_path):
        # 1. 鏂囦欢瀛樺湪锛岃拷鍔犳ā寮?
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        existing_headers = [cell.value for cell in sheet[1]]
        row_data = [item.get(header) for header in existing_headers]

        sheet.append(row_data)
        
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = list(item.keys())
        
        sheet.append(headers)
        row_data = [item.get(header) for header in headers]
        sheet.append(row_data)

    wb.save(file_path)
    wb.close()

import json
import re

def parse_llm_answer(llm_output: str) -> dict:
    """
    瑙ｆ瀽澶фā鍨嬭緭鍑虹殑鏂囨湰锛屾彁鍙栧垎鏋愯繃绋嬪拰 JSON 鏍囨敞缁撴灉銆?
    
    :param llm_output: 澶фā鍨嬭繑鍥炵殑鍘熷瀛楃涓?
    :return: 鍖呭惈鎵€鏈夋爣娉ㄥ瓧娈典互鍙?'Analysis' 瀛楁鐨勫瓧鍏搞€傚鏋滆В鏋愬け璐ワ紝杩斿洖甯︽湁閿欒鏍囪鐨勫瓧鍏搞€?
    """
    parsed_data = {}
    
    # 1. 鎻愬彇 <analysis> 鏍囩涓殑鍐呭 (闈炶椽濠尮閰嶏紝澶勭悊璺ㄨ)
    analysis_match = re.search(r"<analysis>(.*?)</analysis>", llm_output, re.DOTALL | re.IGNORECASE)
    if analysis_match:
        parsed_data['Analysis'] = analysis_match.group(1).strip()
    else:
        parsed_data['Analysis'] = "鏈彁鍙栧埌鍒嗘瀽杩囩▼"

    # 2. 鎻愬彇 JSON 閮ㄥ垎
    # 瀵绘壘绗竴涓?'{' 鍜屾渶鍚庝竴涓?'}' 涔嬮棿鐨勬墍鏈夊唴瀹癸紝杩欐牱鍙互鑷姩蹇界暐妯″瀷涔卞姞鐨?```json 鎴栧叾浠栧簾璇?
    json_match = re.search(r"(\{.*\})", llm_output, re.DOTALL)
    
    if json_match:
        json_str = json_match.group(1)
        try:
            # 灏濊瘯瑙ｆ瀽 JSON
            json_dict = json.loads(json_str)
            # 灏嗚В鏋愬嚭鏉ョ殑 9 涓瓧娈垫洿鏂板埌鎴戜滑鐨勮繑鍥炲瓧鍏镐腑
            parsed_data.update(json_dict)
        except json.JSONDecodeError as e:
            print(f"JSON 瑙ｆ瀽閿欒: {e}")
            print(f"灏濊瘯瑙ｆ瀽鐨勫瓧绗︿覆: \n{json_str}")
            parsed_data['Parse_Error'] = True
    else:
        print("鏈湪妯″瀷杈撳嚭涓壘鍒?JSON 鏍煎紡鏁版嵁")
        parsed_data['Parse_Error'] = True
        
    return parsed_data


if __name__ == "__main__":
    content = "瑙嗛涓煡璇㈠埌鐨勫鐢熷嚭鐢熷勾浠芥渶澶х浉宸嚑骞达紵"
    content = render_template(
        "templates/property.md",
        {
            "question": content
        }
    )
    model_answ = LLM_response(content, model='gemini-2.5-pro')
    print(f"{model_answ}")
    parsed_answer = parse_llm_answer(model_answ)

    print(f"parsed answer:{parsed_answer}")


